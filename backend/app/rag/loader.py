"""
Document loader — dispatches to the appropriate document loader by file extension.
Supports: PDF, DOCX, TXT, Markdown, CSV, XLS, XLSX, HTML, JSON, and text-based formats.
"""
import os
import re
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    CSVLoader,
    BSHTMLLoader,
    Docx2txtLoader,
)


def _clean_text(text: str) -> str:
    """Clean unprintable / control characters while preserving formatting."""
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", text)
    return text.strip()


def _load_excel(file_path: str) -> list[Document]:
    """Parse Excel files (.xlsx, .xls) into textual document rows."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        rows_text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if row_vals:
                    sheet_rows.append(" | ".join(row_vals))
            if sheet_rows:
                rows_text.append(f"Sheet: {sheet_name}\n" + "\n".join(sheet_rows))
        
        full_content = "\n\n".join(rows_text)
        return [Document(page_content=full_content, metadata={"source": file_path})]
    except Exception:
        # Fallback to TextLoader or basic file read
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return [Document(page_content=f.read(), metadata={"source": file_path})]
        except Exception:
            return []


def load_document(file_path: str) -> list[Document]:
    """
    Load a document from disk and return a list of LangChain Document objects.

    Supported formats: .pdf, .docx, .doc, .txt, .md, .csv, .xlsx, .xls, .html, .json, .log, .rst
    """
    ext = os.path.splitext(file_path)[1].lower()
    documents: list[Document] = []

    if ext == ".pdf":
        loader = PyPDFLoader(file_path)
        documents = loader.load()
    elif ext in [".docx", ".doc"]:
        try:
            loader = Docx2txtLoader(file_path)
            documents = loader.load()
        except Exception:
            loader = TextLoader(file_path, encoding="utf-8")
            documents = loader.load()
    elif ext in [".xlsx", ".xls"]:
        documents = _load_excel(file_path)
    elif ext == ".csv":
        try:
            loader = CSVLoader(file_path, encoding="utf-8")
            documents = loader.load()
        except Exception:
            loader = CSVLoader(file_path, encoding="latin-1")
            documents = loader.load()
    elif ext in [".html", ".htm"]:
        try:
            loader = BSHTMLLoader(file_path, open_encoding="utf-8")
            documents = loader.load()
        except Exception:
            loader = BSHTMLLoader(file_path, open_encoding="latin-1")
            documents = loader.load()
    elif ext == ".json":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            documents = [Document(page_content=content, metadata={"source": file_path})]
        except Exception:
            loader = TextLoader(file_path, encoding="latin-1")
            documents = loader.load()
    elif ext in [".txt", ".md", ".log", ".rst", ".yaml", ".yml"]:
        try:
            loader = TextLoader(file_path, encoding="utf-8")
            documents = loader.load()
        except UnicodeDecodeError:
            loader = TextLoader(file_path, encoding="latin-1")
            documents = loader.load()
    else:
        try:
            loader = TextLoader(file_path, encoding="utf-8")
            documents = loader.load()
        except Exception:
            loader = TextLoader(file_path, encoding="latin-1")
            documents = loader.load()

    # Clean text content and ensure metadata source
    cleaned_docs = []
    for doc in documents:
        cleaned_content = _clean_text(doc.page_content)
        if cleaned_content:
            doc.page_content = cleaned_content
            doc.metadata["source"] = file_path
            cleaned_docs.append(doc)

    return cleaned_docs
